"""
UTILIDADES PARA EXPORTACION DE REPORTES
Modulo con funciones para exportar datos a PDF y Excel
"""


from weasyprint import HTML, CSS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.templatetags.static import static
from django.utils import timezone
from datetime import datetime
from django.conf import settings


# ====== EXPORTACION A PDF ======
def exportar_pdf(request, titulo, columnas, datos, nombre_archivo):  #  'título' -> 'titulo' (sin tilde)
    logo_url = request.build_absolute_uri(static('img/gym.jpeg'))  #  'logotipo_url' -> 'logo_url'

    # Crear contexto para el template
    contexto = {
        'titulo': titulo,
        'columnas': columnas,
        'datos': datos,
        'logo_url': logo_url,   #  ahora coincide con la variable definida arriba
        'now': timezone.now(),
    }

    html_string = render_to_string('reportes/reporte_pdf.html', contexto)

    html_object = HTML(string=html_string, base_url=request.build_absolute_uri('/'))  # base_url correcta

    pdf_bytes = html_object.write_pdf()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'

    return response

# ====== EXPORTACION A EXCEL ======
def exportar_excel(titulo, columnas, datos, nombre_archivo):
    """
    FUNCIÓN PARA EXPORTAR DATOS A EXCEL USANDO OPENPYXL
    
    Args:
        titulo: Titulo del reporte
        columnas: Lista de nombres de columnas
        datos: Lista de tuplas o diccionarios con los datos
        nombre_archivo: Nombre del archivo Excel a descargar
    
    Returns:
        HttpResponse con el archivo Excel generado
    """
    
    # Crear un nuevo libro de Excel
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte"
    
    # Configurar estilos para el título
    title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    # Agregar titulo
    worksheet.merge_cells('A1:' + chr(64 + len(columnas)) + '1')
    titulo_cell = worksheet['A1']
    titulo_cell.value = titulo
    titulo_cell.font = title_font
    titulo_cell.fill = title_fill
    titulo_cell.alignment = title_alignment
    worksheet.row_dimensions[1].height = 25
    
    # Configurar estilos para los encabezados
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Agregar encabezados de columnas
    for col_num, columna in enumerate(columnas, 1):
        cell = worksheet.cell(row=3, column=col_num)
        cell.value = columna
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    worksheet.row_dimensions[3].height = 20
    
    # Configurar estilos para los datos
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Agregar datos al Excel
    data_fill_alternated = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    for row_num, fila in enumerate(datos, 4):
        # Convertir diccionario a tupla si es necesario
        if isinstance(fila, dict):
            valores = [fila.get(col.lower().replace(' ', '_'), '') for col in columnas]
        else:
            valores = fila
        
        # Llenar las celdas con datos
        for col_num, valor in enumerate(valores, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = valor
            cell.alignment = data_alignment
            cell.border = data_border
            
            # Colorear filas alternas para mejor legibilidad
            if (row_num - 4) % 2 == 0:
                cell.fill = data_fill_alternated
    
    # Ajustar ancho de columnas automaticamente
    for col_num, columna in enumerate(columnas, 1):
        max_length = len(str(columna))
        column_letter = chr(64 + col_num)
        
        for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        
        worksheet.column_dimensions[column_letter].width = max_length + 2
    
    # Crear respuesta HTTP con el Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'
    
    # Guardar el libro en la respuesta
    workbook.save(response)
    
    return response


# ====== GOOGLE FORMS API FUNCTIONS ======

import os
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow

SCOPES = ['https://www.googleapis.com/auth/forms.body', 'https://www.googleapis.com/auth/forms.responses.readonly']

def get_google_forms_service():
    """
    Obtiene el servicio de Google Forms API.
    Requiere credenciales OAuth 2.0.
    """
    creds = None
    # El archivo token.json almacena los tokens de acceso y refresco del usuario
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    # Si no hay credenciales válidas disponibles, permite al usuario iniciar sesión
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            # Usar puerto fijo permite que el redirect URI coincida con el registrado en Google Cloud
            creds = flow.run_local_server(port=8000)
        # Guarda las credenciales para la próxima ejecución
        with open('token.json', 'w') as token:
            token.write(creds.to_json())

    service = build('forms', 'v1', credentials=creds)
    return service

def crear_formulario_google(titulo, descripcion=""):
    """
    Crea un nuevo formulario en Google Forms.
    """
    service = get_google_forms_service()
    
    form = {
        "info": {
            "title": titulo,
            "description": descripcion,
        }
    }
    
    result = service.forms().create(body=form).execute()
    return result

def actualizar_formulario_google(form_id, updates):
    """
    Actualiza un formulario existente en Google Forms.
    updates debe ser un diccionario con los cambios.
    """
    service = get_google_forms_service()
    
    request = service.forms().batchUpdate(formId=form_id, body=updates)
    response = request.execute()
    return response

def obtener_formulario_google(form_id):
    """
    Obtiene la información de un formulario.
    """
    service = get_google_forms_service()
    
    result = service.forms().get(formId=form_id).execute()
    return result

def agregar_preguntas_a_formulario(form_id, formset):
    """
    Agrega preguntas al formulario de Google Forms.
    """
    service = get_google_forms_service()
    
    requests = []
    for form in formset:
        if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
            pregunta = form.cleaned_data['pregunta']
            tipo = form.cleaned_data['tipo']
            opciones = form.cleaned_data.get('opciones')
            requerida = form.cleaned_data.get('requerida', False)
            
            # Mapear tipos de Django a tipos de Google Forms
            tipo_mapping = {
                'short_answer': 'SHORT_ANSWER',
                'paragraph': 'PARAGRAPH',
                'multiple_choice': 'RADIO',
                'check_boxes': 'CHECKBOX',
                'dropdown': 'DROP_DOWN',
                'date': 'DATE',
                'time': 'TIME',
            }
            
            item = {
                'title': pregunta,
                'questionItem': {
                    'question': {
                        'required': requerida,
                    }
                }
            }
            
            if tipo in ['multiple_choice', 'check_boxes', 'dropdown']:
                if opciones:
                    choices = [{'value': op} for op in opciones]
                    item['questionItem']['question']['choiceQuestion'] = {
                        'type': tipo_mapping[tipo],
                        'options': choices
                    }
                else:
                    # Si no hay opciones, usar texto
                    item['questionItem']['question']['textQuestion'] = {}
            else:
                item['questionItem']['question']['textQuestion'] = {}
            
            requests.append({
                'createItem': {
                    'item': item,
                    'location': {'index': 0}
                }
            })
    
    if requests:
        body = {'requests': requests}
        service.forms().batchUpdate(formId=form_id, body=body).execute()
