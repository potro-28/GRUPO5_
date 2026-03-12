from django.shortcuts import render
from django.views.generic import View
from django.views import View as DjangoView
from django.http import HttpResponse
from gimnasio.models import *
from gimnasio.utils import exportar_pdf, exportar_excel
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ====== VISTAS PARA EXPORTAR REPORTES ======

class ExportarAsistenciaPDF(DjangoView):
    """
    VISTA PARA EXPORTAR CATEGORIAS A PDF
    Obtiene todas las categorías y las exporta en formato PDF
    """
    
    def get(self, request):
        # Obtener todas las categorias 
        asistencia = Asistencia.objects.all()
        
        # Definir las columnas que se muestran en el reporte
        columnas = ['ID','Fecha de asistencia','Hora de ingreso','Hora de salida','Membresia','Documento de usuario']
        
        # Preparar los datos en formato de tuplas
        datos = [
            (a.id,a.fecha_asistencia,a.hora_ingreso,a.hora_salida,(a.fk_membresia.fk_usuario.nombre_usuario,a.fk_membresia.fk_usuario.apellido_usuario),a.fk_membresia.fk_usuario.documento)
            for a in asistencia
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Reporte_Asistencia_{datetime.now().strftime("%d_%m_%Y")}'
        
        # Llamar funcion de exportacion a PDF
        return exportar_pdf(
            titulo='REPORTE DE ASISTENCIAS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )


class ExportarAsistenciaExcel(DjangoView):
    """
    VISTA PARA EXPORTAR CATEGORIAS A EXCEL
    Obtiene todas las categorias y las exporta en formato Excel
    """
    
    def get(self, request):
        # Obtener todas las categorias 
        asistencia = Asistencia.objects.all()
        
        # Definir las columnas que se muestran en el reporte
        columnas = ['ID', 'Fecha de asistencia', 'Hora de ingreso', 'Hora de salida', 'Membresia','Documento de usuario']
        
        # Preparar los datos en formato de tuplas
        datos = [
            (a.id,a.fecha_asistencia,a.hora_ingreso,a.hora_salida,(a.fk_membresia.fk_usuario.nombre_usuario,a.fk_membresia.fk_usuario.apellido_usuario),a.fk_membresia.fk_usuario.documento)
            for a in asistencia
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Reporte_Asistencia_{datetime.now().strftime("%d_%m_%Y")}'
        
        # Llamar funcion de exportacion a PDF
        return exportar_excel(
            titulo='REPORTE DE ASISTENCIAS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )


# ====== VISTAS PARA EXPORTAR REPORTES MEMBRESIA======

class ExportarMembresiaPDF(DjangoView):
    """
    VISTA PARA EXPORTAR CATEGORIAS A PDF
    Obtiene todas las categorías y las exporta en formato PDF
    """
    
    def get(self, request):
        # Obtener todas las categorias 
        membresia = Membresia.objects.all()
        
        # Definir las columnas que se 
        # 
        # en el reporte
        columnas = ['ID','Fecha de inicio','Fecha de fin','Membresia','Documento de usuario']
        
        # Preparar los datos en formato de tuplas
        datos = [
            (m.id,m.fecha_inicio, m.fecha_fin,(m.fk_usuario.nombre_usuario,m.fk_usuario.apellido_usuario),m.fk_usuario.documento)
            for m in membresia
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Reporte_membresia_{datetime.now().strftime("%d_%m_%Y")}'
        
        # Llamar funcion de exportacion a PDF
        return exportar_pdf(
            titulo='REPORTE DE MEMBRESIAS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )


class ExportarMembresiaExcel(DjangoView):
    """
    VISTA PARA EXPORTAR CATEGORIAS A EXCEL
    Obtiene todas las categorias y las exporta en formato Excel
    """
    
    def get(self, request):
        # Obtener todas las categorias 
        membresia = Membresia.objects.all()
        
        # Definir las columnas que se 
        # 
        # en el reporte
        columnas = ['ID','Fecha de inicio','Fecha de fin','Membresia','Documento de usuario']
        
        # Preparar los datos en formato de tuplas
        datos = [
            (m.id,m.fecha_inicio, m.fecha_fin,(m.fk_usuario.nombre_usuario,m.fk_usuario.apellido_usuario),m.fk_usuario.documento)
            for m in membresia
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Reporte_membresia_{datetime.now().strftime("%d_%m_%Y")}'    
        
        # Llamar funcion de exportacion a PDF
        return exportar_excel(
            titulo='REPORTE DE MEMBRESIAS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )



























































































































































































































































































































































































































































































class ExportarSancionPDF(DjangoView):
    
    def get(self, request):
        sanciones = Sancion.objects.all()

        columnas = ['Fecha', 'Nombre', 'Cédula', 'Motivo de sanción']

        datos = [
            (
                s.fecha_inicio,
                s.fk_usuario.nombre_usuario,
                s.fk_usuario.documento,
                s.motivo_sancion
            )
            for s in sanciones
        ]

        nombre_archivo = f'Reporte_sanciones_{datetime.now().strftime("%d_%m_%Y")}'

        return exportar_pdf(
            titulo='REPORTE DE SANCIONES',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )
class ExportarSancionExcel(DjangoView):
    
    def get(self, request):

        sanciones = Sancion.objects.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Sanciones"

        # Encabezados
        columnas = ['Fecha', 'Nombre', 'Cédula', 'Motivo de sanción']
        ws.append(columnas)

        # Estilo encabezados
        for col in ws[1]:
            col.font = Font(bold=True, color="FFFFFF")
            col.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            col.alignment = Alignment(horizontal="center")

        # Datos
        for s in sanciones:
            ws.append([
                s.fecha_inicio.strftime("%d/%m/%Y"),
                s.fk_usuario.nombre_usuario,
                s.fk_usuario.documento,
                s.motivo_sancion
            ])

        # Ajustar ancho columnas automáticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[column_letter].width = max_length + 2

        nombre_archivo = f'Reporte_sanciones_{datetime.now().strftime("%d_%m_%Y")}.xlsx'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

        wb.save(response)

        return response